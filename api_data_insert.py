# -*- coding: utf-8 -*- 
from openpyxl import load_workbook
from pymongo import MongoClient

client = MongoClient('localhost', 27017)
db = client.dbweather

# data_only=True로 해줘야 수식이 아닌 값으로 받아온다.
load_wb = load_workbook('기상청41_단기예보 조회서비스_오픈API활용가이드_격자_위경도(20210401).xlsx', data_only=True)

# 시트 이름으로 불러오기
load_ws = load_wb['최종 업데이트 파일_20210401']

code = load_ws.cell(3775, 2).value  # 행정구역코드
print(code)

for i in range(2, 3776):
    code = load_ws.cell(i, 2).value
    sido = load_ws.cell(i, 3).value
    sigungu = load_ws.cell(i, 4).value
    village = load_ws.cell(i, 5).value
    x = load_ws.cell(i, 6).value
    y = load_ws.cell(i, 7).value
    
    if sigungu and village:
        data = ({'code': code, 'sido': sido, 'sigungu': sigungu, 'village': village, 'x': x, 'y': y})
        # print(data)
        # db.grid.insert_one(data)
        
# db.clothes.insert_one({'low_TMP': -100, 'high_TMP': 5, 'clothes': '패딩, 목도리, 장갑, 기모바지', 'msg': '추운 겨울 날씨예요~ 최대한 따뜻하게 입으세요!', 'img': 'r4'})
# db.clothes.insert_one({'low_TMP': 6, 'high_TMP': 9, 'clothes': '겨울코트, 경량패딩, 가죽자켓, 발열 내의', 'msg': '쌀쌀해요~ 두꺼운 외투를 입으세요!', 'img': 'r8'})
# db.clothes.insert_one({'low_TMP': 10, 'high_TMP': 11, 'clothes': '트렌치코트, 항공점퍼, 얇은 코트', 'msg': '겹겹이 레이어드 코디!', 'img': 'r11'})
# db.clothes.insert_one({'low_TMP': 12, 'high_TMP': 16, 'clothes': '자켓, 두꺼운 가디건, 따뜻한 니트, 간절기 외투', 'msg': '얇은 외투를 챙기세요!', 'img': 'r16'})
# db.clothes.insert_one({'low_TMP': 17, 'high_TMP': 19, 'clothes': '니트&가디건, 후드티&맨투맨, 얇은 자켓, 면바지, 슬랙스', 'msg': '두꺼운 외투는 안녕!', 'img': 'r19'})
# db.clothes.insert_one({'low_TMP': 20, 'high_TMP': 22, 'clothes': '긴팔티&후드티, 셔츠, 면바지&슬랙스, 원피스', 'msg': '얆은 옷!', 'img': 'r22'})
# db.clothes.insert_one({'low_TMP': 23, 'high_TMP': 26, 'clothes': '얇은 긴팔, 얇은 셔츠, 반팔, 면바지&반바지', 'msg': '여름이에요! 반팔을 꺼낼 시간~', 'img': 'r27'})
# db.clothes.insert_one({'low_TMP': 27, 'high_TMP': 100, 'clothes': '민소매, 반팔, 반바지, 린넨', 'msg': '한여름!! 얇고 시원하게!', 'img': 'r28'})

